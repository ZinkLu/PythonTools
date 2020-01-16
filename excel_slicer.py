# AUTHOR ZinkLu
# DATE 2019-12-24
import os

from openpyxl import Workbook, load_workbook
from typing import List, AnyStr


class ExcelSlicer:

    def __init__(self, path, pace, save_to="", processor=None):
        self.path = path
        self.pace = pace
        self.save_to = save_to
        self.cursor = 0

        self.template_workbook = self._get_read_only_template()
        self.template_sheet = self.template_workbook.active
        self.template_name = os.path.split(path)[-1].split(".")[0]
        self.template_head = self._get_template_head()

        self.rows = self.template_sheet.rows
        self.processor = processor
        next(self.rows)  # 数据去头

    def _get_read_only_template(self) -> Workbook:
        template = load_workbook(self.path, read_only=True)
        return template

    def _get_template_head(self) -> List[AnyStr]:
        head = next(self.template_sheet.rows)
        return [h.value for h in head]

    def make_template_workbook(self) -> Workbook:
        wb = Workbook()
        sheet = wb.active
        sheet.title = self.template_sheet.title
        sheet.append(self.template_head)
        return wb

    def pace_loopper(self, workbook: Workbook):
        sheet = workbook.active
        for _ in range(self.cursor, self.cursor + self.pace):
            try:
                row = next(self.rows)
                value = [r.value for r in row]
                if self.processor is not None:
                    value = self.processor(value)
                # value[0] = value[1]
                sheet.append(value)
            except StopIteration as e:
                self.cursor += _
                raise e

        self.cursor += self.pace
        return workbook

    def slice_it(self):
        n = 0
        while True:
            workbook = self.make_template_workbook()
            file_path = os.path.join(self.save_to, "会员信息创建" + self.template_name+f"_{n}.xlsx")
            try:
                self.pace_loopper(workbook)
            except StopIteration:
                workbook.save(file_path)
                print(f"saving {file_path}")
                n += 1
                return
            workbook.save(file_path)
            print(f"saving {file_path}")
            n += 1
